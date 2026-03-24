"""
Servicio de exportación a Excel y PDF.
Plantillas con colores institucionales UPDS y orientación vertical.
"""

import io
from datetime import datetime
from collections import defaultdict

from app.services.analizador import sort_key_grupo


TURNOS = {
    "M": "Mañana",
    "T": "Tarde",
    "N": "Noche",
}

# ── Paleta institucional UPDS ──
UPDS_NAVY = "0F1D42"
UPDS_NAVY_LIGHT = "1A2B5F"
UPDS_NAVY_MID = "243A7A"
UPDS_CELESTE = "4A90D9"
UPDS_CELESTE_LIGHT = "6BB0F0"
UPDS_CELESTE_PALE = "E8F2FC"
UPDS_ICE = "F7F9FC"
UPDS_FOG = "EDF1F7"
UPDS_STEEL = "94A3B8"
UPDS_GRAPHITE = "475569"
UPDS_WHITE = "FFFFFF"
ACCENT_ROSE = "F43F5E"
ACCENT_ROSE_PALE = "FFE4E6"


# ═══════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════

def exportar_excel(stats: dict, configs_grupos: dict) -> io.BytesIO:
    """Genera el reporte Excel con colores institucionales UPDS."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Estilos ──
    logo_font = Font(name="Arial", bold=True, size=22, color=UPDS_NAVY)
    subtitulo_logo_font = Font(name="Arial", bold=True, size=10, color=UPDS_CELESTE)
    titulo_font = Font(name="Arial", bold=True, size=13, color=UPDS_NAVY)
    header_font = Font(name="Arial", bold=True, size=10, color=UPDS_WHITE)
    header_fill = PatternFill("solid", fgColor=UPDS_NAVY)
    subheader_fill = PatternFill("solid", fgColor=UPDS_NAVY_MID)
    data_font = Font(name="Arial", size=10)
    data_font_bold = Font(name="Arial", size=10, bold=True)
    repitente_fill = PatternFill("solid", fgColor=ACCENT_ROSE_PALE)
    celeste_pale_fill = PatternFill("solid", fgColor=UPDS_CELESTE_PALE)
    ice_fill = PatternFill("solid", fgColor=UPDS_ICE)
    fog_fill = PatternFill("solid", fgColor=UPDS_FOG)
    border = Border(
        left=Side(style="thin", color=UPDS_FOG),
        right=Side(style="thin", color=UPDS_FOG),
        top=Side(style="thin", color=UPDS_FOG),
        bottom=Side(style="thin", color=UPDS_FOG),
    )
    border_header = Border(
        left=Side(style="thin", color=UPDS_NAVY_LIGHT),
        right=Side(style="thin", color=UPDS_NAVY_LIGHT),
        top=Side(style="thin", color=UPDS_NAVY_LIGHT),
        bottom=Side(style="thin", color=UPDS_NAVY_LIGHT),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _add_logo_header(ws, last_col="E"):
        """Inserta encabezado institucional UPDS en las primeras filas."""
        ws.merge_cells(f"A1:{last_col}1")
        cell = ws.cell(row=1, column=1, value="UPDS")
        cell.font = logo_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=UPDS_WHITE)

        ws.merge_cells(f"A2:{last_col}2")
        cell2 = ws.cell(row=2, column=1, value="Universidad Privada Domingo Savio")
        cell2.font = subtitulo_logo_font
        cell2.alignment = Alignment(horizontal="center", vertical="center")

        # Línea decorativa celeste
        for col_idx in range(1, _col_num(last_col) + 1):
            c = ws.cell(row=3, column=col_idx)
            c.fill = PatternFill("solid", fgColor=UPDS_CELESTE)
            c.border = Border(bottom=Side(style="medium", color=UPDS_NAVY))

        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 4

    def _col_num(letter):
        result = 0
        for ch in letter:
            result = result * 26 + (ord(ch.upper()) - ord('A') + 1)
        return result

    def apply_header(ws, row, cols):
        for col_idx, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border_header

    def apply_data(ws, row, cols, fill=None, bold=False):
        for col_idx, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font_bold if bold else data_font
            cell.border = border
            if fill:
                cell.fill = fill
            elif row % 2 == 0:
                cell.fill = ice_fill
            if isinstance(val, (int, float)):
                cell.alignment = center
            elif col_idx == 1:
                cell.alignment = left_wrap

    # ════════════════════════════════════════
    # Hoja 1: Resumen
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen"
    _add_logo_header(ws1, "E")

    ws1.merge_cells("A5:E5")
    ws1.cell(row=5, column=1, value="REPORTE DE MATRICULADOS - MEDICINA").font = titulo_font
    ws1.cell(row=5, column=1).alignment = Alignment(horizontal="center")

    ws1.merge_cells("A6:E6")
    ws1.cell(row=6, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(name="Arial", size=9, color=UPDS_GRAPHITE)
    ws1.cell(row=6, column=1).alignment = Alignment(horizontal="center")

    # Info cards
    row = 8
    ws1.merge_cells(f"A{row}:B{row}")
    c = ws1.cell(row=row, column=1, value=f"  Total estudiantes únicos: {stats['total_estudiantes_unicos']}")
    c.font = Font(name="Arial", bold=True, size=11, color=UPDS_NAVY)
    c.fill = celeste_pale_fill

    ws1.merge_cells(f"D{row}:E{row}")
    c2 = ws1.cell(row=row, column=4, value=f"  Repitentes detectados: {len(stats['repitentes'])}")
    c2.font = Font(name="Arial", bold=True, size=11, color=ACCENT_ROSE)
    c2.fill = PatternFill("solid", fgColor=ACCENT_ROSE_PALE)

    row = 10
    apply_header(ws1, row, ["Semestre", "Únicos", "Regulares", "Repitentes", "% Repitentes"])
    row += 1
    for sem in sorted(stats["estudiantes_por_semestre"].keys()):
        unicos = len(stats["estudiantes_por_semestre"][sem])
        rep = len(stats["repitentes_por_semestre"].get(sem, set()))
        reg = unicos - rep
        pct = round(rep / unicos * 100, 1) if unicos > 0 else 0
        fill = repitente_fill if rep > 0 else None
        apply_data(ws1, row, [f"Semestre {sem}", unicos, reg, rep, f"{pct}%"], fill)
        row += 1

    for col in ["A", "B", "C", "D", "E"]:
        ws1.column_dimensions[col].width = 18
    ws1.sheet_properties.pageSetUpPr = ws1.sheet_properties.pageSetUpPr or None
    ws1.page_setup.orientation = "portrait"

    # ════════════════════════════════════════
    # Hoja 2: Grupos por Semestre
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("Grupos por Semestre")
    _add_logo_header(ws2, "F")

    ws2.merge_cells("A5:F5")
    ws2.cell(row=5, column=1, value="DETALLE POR GRUPO Y SEMESTRE").font = titulo_font
    ws2.cell(row=5, column=1).alignment = Alignment(horizontal="center")
    row = 7

    for sem in sorted(stats["estudiantes_por_semestre"].keys()):
        if sem in configs_grupos:
            _, grupo_a_letras, grupo_info = configs_grupos[sem]
        else:
            grupo_a_letras = {}
            grupo_info = {}

        total_sem = len(stats["estudiantes_por_semestre"][sem])
        rep_sem = len(stats["repitentes_por_semestre"].get(sem, set()))

        # Subtítulo del semestre con fondo celeste
        ws2.merge_cells(f"A{row}:F{row}")
        c = ws2.cell(row=row, column=1,
                      value=f"  SEMESTRE {sem} — {total_sem} únicos ({total_sem - rep_sem} regulares + {rep_sem} repitentes)")
        c.font = Font(name="Arial", bold=True, size=11, color=UPDS_WHITE)
        c.fill = PatternFill("solid", fgColor=UPDS_NAVY_MID)
        c.alignment = Alignment(vertical="center")
        ws2.row_dimensions[row].height = 26
        row += 1

        apply_header(ws2, row, ["Grupo", "Turno", "Letras", "Matriculados", "Repitentes", "Regulares"])
        row += 1

        grupos_del_sem = {}
        for (s, g), ids in stats["grupo_real_semestre"].items():
            if s == sem:
                grupos_del_sem[g] = ids

        grupo_rep_count = stats.get("grupo_repitentes_count", {})

        for grupo in sorted(grupos_del_sem.keys(), key=lambda g: sort_key_grupo(g, grupo_info)):
            ids = grupos_del_sem[grupo]
            count = len(ids)
            info = grupo_info.get(grupo)
            # Usar conteo directo si está disponible (desde BD), sino calcular por IDs
            if (sem, grupo) in grupo_rep_count:
                rep_g = grupo_rep_count[(sem, grupo)]
            else:
                rep_g = sum(1 for eid in ids if eid in stats["repitentes_por_semestre"].get(sem, set()))

            if info:
                letras = grupo_a_letras.get(grupo, [])
                letras_str = "+".join(letras)
                turno = info["turno_nombre"]
            else:
                letras_str = grupo
                turno = "?"

            fill = repitente_fill if rep_g > 0 else None
            apply_data(ws2, row, [grupo, turno, letras_str, count, rep_g, count - rep_g], fill)
            row += 1

        row += 1

    for col, w in [("A", 10), ("B", 12), ("C", 10), ("D", 14), ("E", 13), ("F", 12)]:
        ws2.column_dimensions[col].width = w
    ws2.page_setup.orientation = "portrait"

    # ════════════════════════════════════════
    # Hoja 3: Repitentes
    # ════════════════════════════════════════
    ws3 = wb.create_sheet("Repitentes")
    _add_logo_header(ws3, "E")

    ws3.merge_cells("A5:E5")
    ws3.cell(row=5, column=1, value=f"ESTUDIANTES REPITENTES ({len(stats['repitentes'])})").font = titulo_font
    ws3.cell(row=5, column=1).alignment = Alignment(horizontal="center")

    row = 7
    apply_header(ws3, row, ["ID", "Nombre", "Semestre Principal", "Semestres donde repite", "Todos los semestres"])
    row += 1

    for est in stats["repitentes"]:
        rep_str = ", ".join(f"S{s}" for s in est["semestres_donde_repite"])
        all_str = ", ".join(f"S{s}" for s in est["todos_los_semestres"])
        apply_data(ws3, row, [est["id"], est["nombre"], est["semestre_principal"], rep_str, all_str], repitente_fill)
        row += 1

    for col, w in [("A", 22), ("B", 48), ("C", 20), ("D", 28), ("E", 22)]:
        ws3.column_dimensions[col].width = w
    ws3.page_setup.orientation = "portrait"

    # ════════════════════════════════════════
    # Hoja 4: Origen Repitentes
    # ════════════════════════════════════════
    ws4 = wb.create_sheet("Origen Repitentes")
    _add_logo_header(ws4, "D")

    ws4.merge_cells("A5:D5")
    ws4.cell(row=5, column=1, value="ORIGEN DE REPITENTES POR SEMESTRE").font = titulo_font
    ws4.cell(row=5, column=1).alignment = Alignment(horizontal="center")

    row = 7
    apply_header(ws4, row, ["Semestre donde repite", "Total repitentes", "Vienen del semestre", "Cantidad"])
    row += 1

    for sem in sorted(stats["repitentes_por_semestre"].keys()):
        ids = stats["repitentes_por_semestre"][sem]
        origenes = defaultdict(int)
        for eid in ids:
            origenes[stats["est_sem_principal"][eid]] += 1
        first = True
        for s_o in sorted(origenes.keys()):
            if first:
                apply_data(ws4, row, [f"Semestre {sem}", len(ids), f"Semestre {s_o}", origenes[s_o]], bold=True)
                first = False
            else:
                apply_data(ws4, row, ["", "", f"Semestre {s_o}", origenes[s_o]])
            row += 1
        row += 1

    for col, w in [("A", 22), ("B", 18), ("C", 22), ("D", 12)]:
        ws4.column_dimensions[col].width = w
    ws4.page_setup.orientation = "portrait"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════
#  PDF
# ═══════════════════════════════════════════════════════════

def _draw_upds_logo(canvas, x, y, size=18):
    """Dibuja el texto 'UPDS' como logo institucional."""
    canvas.saveState()
    # Fondo navy para las letras
    text_width = canvas.stringWidth("UPDS", "Helvetica-Bold", size)
    canvas.setFillColor(_hex("#0F1D42"))
    canvas.setFont("Helvetica-Bold", size)
    canvas.drawString(x, y, "UPDS")
    canvas.restoreState()


def _hex(color_str):
    """Convierte hex string a reportlab color."""
    from reportlab.lib.colors import HexColor
    return HexColor(color_str)


def _header_footer(canvas, doc):
    """Dibuja encabezado y pie de página institucional en cada página."""
    from reportlab.lib.units import inch, cm
    from reportlab.lib.colors import HexColor

    width, height = doc.pagesize
    navy = HexColor("#0F1D42")
    celeste = HexColor("#4A90D9")
    celeste_pale = HexColor("#E8F2FC")
    graphite = HexColor("#475569")

    canvas.saveState()

    # ── Encabezado ──
    # Barra superior navy
    canvas.setFillColor(navy)
    canvas.rect(0, height - 52, width, 52, fill=1, stroke=0)

    # Logo UPDS en blanco
    canvas.setFillColor(HexColor("#FFFFFF"))
    canvas.setFont("Helvetica-Bold", 22)
    canvas.drawString(doc.leftMargin, height - 38, "UPDS")

    # Subtítulo
    canvas.setFont("Helvetica", 8)
    canvas.drawString(doc.leftMargin + 68, height - 36, "Universidad Privada Domingo Savio")

    # Texto derecho
    canvas.setFont("Helvetica", 7)
    canvas.drawRightString(width - doc.rightMargin, height - 32, "Sistema de Conteo de Registros")
    canvas.drawRightString(width - doc.rightMargin, height - 42, "Carrera de Medicina")

    # Línea celeste debajo del encabezado
    canvas.setStrokeColor(celeste)
    canvas.setLineWidth(2)
    canvas.line(0, height - 54, width, height - 54)

    # ── Pie de página ──
    # Línea separadora
    canvas.setStrokeColor(celeste)
    canvas.setLineWidth(1)
    canvas.line(doc.leftMargin, 36, width - doc.rightMargin, 36)

    # Texto del pie
    canvas.setFillColor(graphite)
    canvas.setFont("Helvetica", 7)
    canvas.drawString(doc.leftMargin, 24,
                      f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  SisCoRe UPDS")
    canvas.drawRightString(width - doc.rightMargin, 24,
                           f"Página {doc.page}")

    # Barra inferior navy delgada
    canvas.setFillColor(navy)
    canvas.rect(0, 0, width, 14, fill=1, stroke=0)

    canvas.restoreState()


def exportar_pdf(stats: dict, configs_grupos: dict) -> io.BytesIO:
    """Genera el reporte PDF vertical con branding institucional UPDS."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable,
    )
    from reportlab.lib.units import inch, cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    output = io.BytesIO()

    NAVY = colors.HexColor("#0F1D42")
    NAVY_MID = colors.HexColor("#243A7A")
    CELESTE = colors.HexColor("#4A90D9")
    CELESTE_PALE = colors.HexColor("#E8F2FC")
    ICE = colors.HexColor("#F7F9FC")
    FOG = colors.HexColor("#EDF1F7")
    GRAPHITE = colors.HexColor("#475569")
    ROSE = colors.HexColor("#F43F5E")
    ROSE_PALE = colors.HexColor("#FFE4E6")

    doc = SimpleDocTemplate(
        output,
        pagesize=letter,  # Vertical (portrait)
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=1.0 * inch,   # Espacio para el encabezado
        bottomMargin=0.7 * inch,  # Espacio para el pie
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="TituloReporte",
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=NAVY,
        spaceAfter=4,
        spaceBefore=8,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="Subtitulo",
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=NAVY,
        spaceAfter=6,
        spaceBefore=14,
    ))
    styles.add(ParagraphStyle(
        name="SubtituloSemestre",
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.white,
        spaceAfter=4,
        spaceBefore=10,
        backColor=NAVY_MID,
        borderPadding=(6, 8, 6, 8),
    ))
    styles.add(ParagraphStyle(
        name="Info",
        fontName="Helvetica",
        fontSize=10,
        textColor=GRAPHITE,
        spaceAfter=2,
    ))
    styles.add(ParagraphStyle(
        name="InfoCenter",
        fontName="Helvetica",
        fontSize=10,
        textColor=GRAPHITE,
        spaceAfter=2,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="CardValue",
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=NAVY,
        alignment=TA_CENTER,
        spaceAfter=0,
    ))
    styles.add(ParagraphStyle(
        name="CardLabel",
        fontName="Helvetica",
        fontSize=8,
        textColor=GRAPHITE,
        alignment=TA_CENTER,
        spaceBefore=0,
    ))
    styles.add(ParagraphStyle(
        name="CellText",
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    ))
    styles.add(ParagraphStyle(
        name="CellTextBold",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
    ))

    story = []

    # Estilos de tabla base
    base_table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, 0), 0.5, NAVY),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, FOG),
        ("LINEAFTER", (0, 0), (-2, -1), 0.3, FOG),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ICE]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]

    # ════════════════════════════════════════
    # Portada
    # ════════════════════════════════════════
    story.append(Spacer(1, 40))
    story.append(Paragraph("REPORTE DE MATRICULADOS", styles["TituloReporte"]))
    story.append(Paragraph("Carrera de Medicina", styles["InfoCenter"]))

    story.append(Spacer(1, 6))
    story.append(HRFlowable(
        width="40%", thickness=2, color=CELESTE,
        spaceAfter=16, spaceBefore=6, hAlign="CENTER",
    ))

    # Cards resumen
    total_unicos = stats['total_estudiantes_unicos']
    total_rep = len(stats['repitentes'])
    card_data = [
        [
            Paragraph(str(total_unicos), styles["CardValue"]),
            Paragraph(str(total_rep), styles["CardValue"]),
        ],
        [
            Paragraph("Estudiantes únicos", styles["CardLabel"]),
            Paragraph("Repitentes detectados", styles["CardLabel"]),
        ],
    ]
    card_table = Table(card_data, colWidths=[200, 200])
    card_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), CELESTE_PALE),
        ("BACKGROUND", (1, 0), (1, -1), ROSE_PALE),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("BOX", (0, 0), (0, -1), 1, CELESTE),
        ("BOX", (1, 0), (1, -1), 1, ROSE),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
    ]))
    story.append(card_table)

    story.append(Spacer(1, 20))

    # Tabla resumen por semestre
    story.append(Paragraph("RESUMEN POR SEMESTRE", styles["Subtitulo"]))

    avail_width = doc.width
    data = [["Semestre", "Únicos", "Regulares", "Repitentes", "% Repitentes"]]
    row_fills = []
    for sem in sorted(stats["estudiantes_por_semestre"].keys()):
        u = len(stats["estudiantes_por_semestre"][sem])
        r = len(stats["repitentes_por_semestre"].get(sem, set()))
        pct = f"{r / u * 100:.1f}%" if u > 0 else "0%"
        data.append([f"Semestre {sem}", str(u), str(u - r), str(r), pct])
        row_fills.append(r > 0)

    col_w = [avail_width * 0.28, avail_width * 0.18, avail_width * 0.18, avail_width * 0.18, avail_width * 0.18]
    t = Table(data, colWidths=col_w)
    style_cmds = list(base_table_style)
    for i, has_rep in enumerate(row_fills):
        if has_rep:
            style_cmds.append(("BACKGROUND", (3, i + 1), (3, i + 1), ROSE_PALE))
    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    # ════════════════════════════════════════
    # Detalle por semestre
    # ════════════════════════════════════════
    for sem in sorted(stats["estudiantes_por_semestre"].keys()):
        story.append(PageBreak())
        total_sem = len(stats["estudiantes_por_semestre"][sem])
        rep_sem = len(stats["repitentes_por_semestre"].get(sem, set()))

        story.append(Paragraph(
            f"SEMESTRE {sem} &mdash; {total_sem} únicos ({total_sem - rep_sem} reg. + {rep_sem} repit.)",
            styles["SubtituloSemestre"]
        ))

        if sem in configs_grupos:
            _, grupo_a_letras, grupo_info = configs_grupos[sem]
        else:
            grupo_a_letras = {}
            grupo_info = {}

        data = [["Grupo", "Turno", "Letras", "Matriculados", "Repitentes", "Regulares"]]
        sem_row_fills = []

        grupos_del_sem = {}
        for (s, g), ids in stats["grupo_real_semestre"].items():
            if s == sem:
                grupos_del_sem[g] = ids

        grupo_rep_count = stats.get("grupo_repitentes_count", {})

        for grupo in sorted(grupos_del_sem.keys(), key=lambda g: sort_key_grupo(g, grupo_info)):
            ids = grupos_del_sem[grupo]
            count = len(ids)
            info = grupo_info.get(grupo)
            # Usar conteo directo si está disponible (desde BD), sino calcular por IDs
            if (sem, grupo) in grupo_rep_count:
                rep_g = grupo_rep_count[(sem, grupo)]
            else:
                rep_g = sum(1 for eid in ids if eid in stats["repitentes_por_semestre"].get(sem, set()))

            turno = info["turno_nombre"] if info else "?"
            letras_str = "+".join(grupo_a_letras.get(grupo, [grupo]))
            data.append([grupo, turno, letras_str, str(count), str(rep_g), str(count - rep_g)])
            sem_row_fills.append(rep_g > 0)

        col_w = [avail_width * 0.12, avail_width * 0.18, avail_width * 0.12,
                 avail_width * 0.20, avail_width * 0.18, avail_width * 0.20]
        t = Table(data, colWidths=col_w)
        style_cmds = list(base_table_style)
        for i, has_rep in enumerate(sem_row_fills):
            if has_rep:
                style_cmds.append(("BACKGROUND", (0, i + 1), (-1, i + 1), ROSE_PALE))
        t.setStyle(TableStyle(style_cmds))
        story.append(t)

    # ════════════════════════════════════════
    # Lista de repitentes
    # ════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph(
        f"ESTUDIANTES REPITENTES ({len(stats['repitentes'])})",
        styles["Subtitulo"]
    ))

    if stats["repitentes"]:
        data = [["ID", "Nombre", "Sem. Principal", "Repite en"]]
        for est in stats["repitentes"]:
            rep_str = ", ".join(f"S{s}" for s in est["semestres_donde_repite"])
            nombre = est["nombre"][:55]
            data.append([
                Paragraph(str(est["id"]), styles["CellText"]),
                Paragraph(nombre, styles["CellText"]),
                str(est["semestre_principal"]),
                rep_str,
            ])

        col_w = [avail_width * 0.22, avail_width * 0.40, avail_width * 0.16, avail_width * 0.22]
        t = Table(data, colWidths=col_w, repeatRows=1)
        style_cmds = list(base_table_style)
        # Fondo rosado suave para todas las filas de repitentes
        style_cmds.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [ROSE_PALE, colors.HexColor("#FFF5F5")]))
        t.setStyle(TableStyle(style_cmds))
        story.append(t)

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════
#  PDF DE LISTAS PARA IMPRESIÓN
# ═══════════════════════════════════════════════════════════

def exportar_listas_pdf(listas: list[dict], periodo_nombre: str) -> io.BytesIO:
    """
    Genera un PDF vertical con listas de estudiantes por grupo,
    pensado para impresión. Una página (o más) por grupo.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable, KeepTogether,
    )
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    output = io.BytesIO()

    NAVY = colors.HexColor("#0F1D42")
    NAVY_MID = colors.HexColor("#243A7A")
    CELESTE = colors.HexColor("#4A90D9")
    CELESTE_PALE = colors.HexColor("#E8F2FC")
    ICE = colors.HexColor("#F7F9FC")
    FOG = colors.HexColor("#EDF1F7")
    GRAPHITE = colors.HexColor("#475569")

    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=1.0 * inch,
        bottomMargin=0.7 * inch,
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="ListaTitulo",
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=NAVY,
        spaceAfter=2,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="ListaSubtitulo",
        fontName="Helvetica",
        fontSize=9,
        textColor=GRAPHITE,
        spaceAfter=8,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="GrupoTitulo",
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.white,
        backColor=NAVY_MID,
        borderPadding=(6, 10, 6, 10),
        spaceAfter=0,
        spaceBefore=0,
    ))
    styles.add(ParagraphStyle(
        name="GrupoInfo",
        fontName="Helvetica",
        fontSize=9,
        textColor=GRAPHITE,
        spaceAfter=6,
        spaceBefore=4,
    ))
    styles.add(ParagraphStyle(
        name="CeldaTexto",
        fontName="Helvetica",
        fontSize=9,
        leading=11,
    ))
    styles.add(ParagraphStyle(
        name="CeldaNum",
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        alignment=TA_CENTER,
    ))

    avail_width = doc.width

    # Estilo base para tablas de lista
    lista_table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),   # Nro centrado
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, FOG),
        ("LINEAFTER", (0, 0), (-2, -1), 0.3, FOG),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ICE]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        # Borde exterior
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]

    story = []

    for idx, lista in enumerate(listas):
        if idx > 0:
            story.append(PageBreak())

        sem = lista["semestre"]
        grp = lista["grupo"]
        turno = lista["turno"]
        estudiantes = lista["estudiantes"]
        rep_count = lista.get("repitentes_count", 0)

        # Encabezado del grupo
        story.append(Paragraph(
            f"LISTA DE ESTUDIANTES &mdash; SEMESTRE {sem}",
            styles["ListaTitulo"],
        ))
        story.append(Paragraph(
            f"{periodo_nombre} &bull; Carrera de Medicina &bull; UPDS",
            styles["ListaSubtitulo"],
        ))

        story.append(Spacer(1, 4))

        # Barra info del grupo
        story.append(Paragraph(
            f"Grupo {grp} &mdash; Turno {turno}",
            styles["GrupoTitulo"],
        ))
        story.append(Paragraph(
            f"Total matriculados: {len(estudiantes)}  |  Repitentes: {rep_count}  |  Regulares: {len(estudiantes) - rep_count}",
            styles["GrupoInfo"],
        ))

        # Tabla de estudiantes
        col_widths = [
            avail_width * 0.08,   # Nro
            avail_width * 0.25,   # ID
            avail_width * 0.47,   # Nombre
            avail_width * 0.20,   # Firma
        ]
        data = [["N°", "Código", "Nombre Completo", "Firma"]]

        for i, est in enumerate(estudiantes, 1):
            data.append([
                str(i),
                Paragraph(str(est["id"]), styles["CeldaTexto"]),
                Paragraph(est["nombre"], styles["CeldaTexto"]),
                "",  # Espacio para firma
            ])

        t = Table(data, colWidths=col_widths, repeatRows=1)
        style_cmds = list(lista_table_style)

        # Línea para firma en la última columna
        for row_idx in range(1, len(data)):
            style_cmds.append(
                ("LINEBELOW", (3, row_idx), (3, row_idx), 0.5, colors.HexColor("#94A3B8"))
            )

        t.setStyle(TableStyle(style_cmds))
        story.append(t)

        # Pie con espacio para firmas de docente
        story.append(Spacer(1, 30))
        firma_data = [
            ["", ""],
            ["_" * 35, "_" * 35],
            ["Firma del Docente", "Firma del Director"],
        ]
        firma_table = Table(firma_data, colWidths=[avail_width * 0.45, avail_width * 0.45])
        firma_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 2), (-1, 2), "Helvetica"),
            ("FONTSIZE", (0, 2), (-1, 2), 8),
            ("TEXTCOLOR", (0, 2), (-1, 2), GRAPHITE),
            ("TOPPADDING", (0, 2), (-1, 2), 4),
        ]))
        story.append(firma_table)

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════
#  EXPORTAR LISTAS POR MATERIA
# ═══════════════════════════════════════════════════════════


def exportar_materias_pdf(materias: list[dict], periodo_nombre: str) -> io.BytesIO:
    """
    Genera un PDF vertical con listas de estudiantes por materia,
    pensado para impresión. Una página (o más) por materia.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak,
    )
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER

    output = io.BytesIO()

    NAVY = colors.HexColor("#0F1D42")
    NAVY_MID = colors.HexColor("#243A7A")
    CELESTE = colors.HexColor("#4A90D9")
    ICE = colors.HexColor("#F7F9FC")
    FOG = colors.HexColor("#EDF1F7")
    GRAPHITE = colors.HexColor("#475569")

    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=1.0 * inch,
        bottomMargin=0.7 * inch,
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="MatTitulo",
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=NAVY,
        spaceAfter=2,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="MatSubtitulo",
        fontName="Helvetica",
        fontSize=9,
        textColor=GRAPHITE,
        spaceAfter=8,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="MatGrupoTitulo",
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.white,
        backColor=NAVY_MID,
        borderPadding=(6, 10, 6, 10),
        spaceAfter=0,
        spaceBefore=0,
    ))
    styles.add(ParagraphStyle(
        name="MatGrupoInfo",
        fontName="Helvetica",
        fontSize=9,
        textColor=GRAPHITE,
        spaceAfter=6,
        spaceBefore=4,
    ))
    styles.add(ParagraphStyle(
        name="MatCeldaTexto",
        fontName="Helvetica",
        fontSize=9,
        leading=11,
    ))

    avail_width = doc.width

    lista_table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, FOG),
        ("LINEAFTER", (0, 0), (-2, -1), 0.3, FOG),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ICE]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]

    story = []

    for idx, materia in enumerate(materias):
        if idx > 0:
            story.append(PageBreak())

        sem = materia["semestre"]
        codigo = materia["codigo"]
        nombre = materia["nombre"]
        letra = materia["letra"]
        grupo = materia["grupo"]
        estudiantes = materia["estudiantes"]

        story.append(Paragraph(
            f"LISTA DE REGISTRADOS &mdash; SEMESTRE {sem}",
            styles["MatTitulo"],
        ))
        story.append(Paragraph(
            f"{periodo_nombre} &bull; Carrera de Medicina &bull; UPDS",
            styles["MatSubtitulo"],
        ))

        story.append(Spacer(1, 4))

        story.append(Paragraph(
            f"{codigo} &mdash; {nombre} (Letra {letra} / Grupo {grupo})",
            styles["MatGrupoTitulo"],
        ))
        story.append(Paragraph(
            f"Total registrados: {len(estudiantes)}",
            styles["MatGrupoInfo"],
        ))

        col_widths = [
            avail_width * 0.08,
            avail_width * 0.25,
            avail_width * 0.47,
            avail_width * 0.20,
        ]
        data = [["N°", "Código", "Nombre Completo", "Firma"]]

        for i, est in enumerate(estudiantes, 1):
            data.append([
                str(i),
                Paragraph(str(est["id"]), styles["MatCeldaTexto"]),
                Paragraph(est["nombre"], styles["MatCeldaTexto"]),
                "",
            ])

        t = Table(data, colWidths=col_widths, repeatRows=1)
        style_cmds = list(lista_table_style)

        for row_idx in range(1, len(data)):
            style_cmds.append(
                ("LINEBELOW", (3, row_idx), (3, row_idx), 0.5, colors.HexColor("#94A3B8"))
            )

        t.setStyle(TableStyle(style_cmds))
        story.append(t)

        story.append(Spacer(1, 30))
        firma_data = [
            ["", ""],
            ["_" * 35, "_" * 35],
            ["Firma del Docente", "Firma del Director"],
        ]
        firma_table = Table(firma_data, colWidths=[avail_width * 0.45, avail_width * 0.45])
        firma_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 2), (-1, 2), "Helvetica"),
            ("FONTSIZE", (0, 2), (-1, 2), 8),
            ("TEXTCOLOR", (0, 2), (-1, 2), GRAPHITE),
            ("TOPPADDING", (0, 2), (-1, 2), 4),
        ]))
        story.append(firma_table)

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════
#  EXPORTAR REPITENTES
# ═══════════════════════════════════════════════════════════

def exportar_repitentes_excel(repitentes: list[dict], periodo_nombre: str) -> io.BytesIO:
    """Genera un Excel con la lista de repitentes filtrada."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Repitentes"

    # Estilos
    logo_font = Font(name="Arial", bold=True, size=22, color=UPDS_NAVY)
    subtitulo_logo_font = Font(name="Arial", bold=True, size=10, color=UPDS_CELESTE)
    titulo_font = Font(name="Arial", bold=True, size=13, color=UPDS_NAVY)
    header_font = Font(name="Arial", bold=True, size=10, color=UPDS_WHITE)
    header_fill = PatternFill("solid", fgColor=UPDS_NAVY)
    data_font = Font(name="Arial", size=10)
    ice_fill = PatternFill("solid", fgColor=UPDS_ICE)
    rose_fill = PatternFill("solid", fgColor=ACCENT_ROSE_PALE)
    border = Border(
        left=Side(style="thin", color=UPDS_FOG),
        right=Side(style="thin", color=UPDS_FOG),
        top=Side(style="thin", color=UPDS_FOG),
        bottom=Side(style="thin", color=UPDS_FOG),
    )
    border_header = Border(
        left=Side(style="thin", color=UPDS_NAVY_LIGHT),
        right=Side(style="thin", color=UPDS_NAVY_LIGHT),
        top=Side(style="thin", color=UPDS_NAVY_LIGHT),
        bottom=Side(style="thin", color=UPDS_NAVY_LIGHT),
    )
    center = Alignment(horizontal="center", vertical="center")

    # Encabezado institucional
    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="UPDS").font = logo_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1, value="Universidad Privada Domingo Savio").font = subtitulo_logo_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    for col_idx in range(1, 6):
        c = ws.cell(row=3, column=col_idx)
        c.fill = PatternFill("solid", fgColor=UPDS_CELESTE)
    ws.row_dimensions[3].height = 4

    ws.merge_cells("A5:E5")
    ws.cell(row=5, column=1, value=f"ESTUDIANTES REPITENTES — {periodo_nombre}").font = titulo_font
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells("A6:E6")
    ws.cell(row=6, column=1, value=f"Total: {len(repitentes)} repitentes  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(name="Arial", size=9, color=UPDS_GRAPHITE)
    ws.cell(row=6, column=1).alignment = Alignment(horizontal="center")

    # Header de tabla
    row = 8
    headers = ["ID Estudiante", "Nombre", "Sem. Principal", "Semestres donde repite", "Todos los semestres"]
    for col_idx, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_header

    # Datos
    row = 9
    for r in repitentes:
        rep_str = ", ".join(f"S{s}" for s in r["semestres_donde_repite"])
        all_str = ", ".join(f"S{s}" for s in r["todos_los_semestres"])
        valores = [r["estudiante_id"], r["nombre"], r["semestre_principal"], rep_str, all_str]
        fill = rose_fill if row % 2 == 0 else ice_fill
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            cell.fill = fill
            if isinstance(val, int):
                cell.alignment = center
        row += 1

    # Anchos
    for col, w in [("A", 22), ("B", 48), ("C", 16), ("D", 28), ("E", 24)]:
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "portrait"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def exportar_repitentes_pdf(repitentes: list[dict], periodo_nombre: str) -> io.BytesIO:
    """Genera un PDF vertical con la lista de repitentes filtrada."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable,
    )
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER

    output = io.BytesIO()

    NAVY = colors.HexColor("#0F1D42")
    CELESTE = colors.HexColor("#4A90D9")
    ICE = colors.HexColor("#F7F9FC")
    FOG = colors.HexColor("#EDF1F7")
    GRAPHITE = colors.HexColor("#475569")

    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=1.0 * inch,
        bottomMargin=0.7 * inch,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="RepTitulo", fontName="Helvetica-Bold", fontSize=16,
        textColor=NAVY, spaceAfter=4, alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="RepSubtitulo", fontName="Helvetica", fontSize=9,
        textColor=GRAPHITE, spaceAfter=8, alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="RepCelda", fontName="Helvetica", fontSize=8, leading=10,
    ))

    avail_width = doc.width
    story = []

    story.append(Spacer(1, 20))
    story.append(Paragraph("ESTUDIANTES REPITENTES", styles["RepTitulo"]))
    story.append(Paragraph(
        f"{periodo_nombre} &bull; Carrera de Medicina &bull; UPDS",
        styles["RepSubtitulo"],
    ))
    story.append(HRFlowable(width="30%", thickness=2, color=CELESTE, spaceAfter=10, spaceBefore=4, hAlign="CENTER"))
    story.append(Paragraph(
        f"Total: <b>{len(repitentes)}</b> repitentes",
        styles["RepSubtitulo"],
    ))
    story.append(Spacer(1, 10))

    # Tabla
    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, FOG),
        ("LINEAFTER", (0, 0), (-2, -1), 0.3, FOG),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ICE]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]

    data = [["N", "ID", "Nombre", "Sem.", "Repite en", "Todos"]]
    for i, r in enumerate(repitentes, 1):
        rep_str = ", ".join(f"S{s}" for s in r["semestres_donde_repite"])
        all_str = ", ".join(f"S{s}" for s in r["todos_los_semestres"])
        data.append([
            str(i),
            Paragraph(str(r["estudiante_id"]), styles["RepCelda"]),
            Paragraph(r["nombre"][:50], styles["RepCelda"]),
            str(r["semestre_principal"]),
            rep_str,
            all_str,
        ])

    col_w = [
        avail_width * 0.05,  # N
        avail_width * 0.18,  # ID
        avail_width * 0.33,  # Nombre
        avail_width * 0.07,  # Sem
        avail_width * 0.18,  # Repite en
        avail_width * 0.19,  # Todos
    ]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle(base_style))
    story.append(t)

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    output.seek(0)
    return output
