"""
Sistema de Análisis de Matriculados - Medicina UPDS v3.0
=========================================================
- Configuración de grupos POR SEMESTRE (cada semestre puede tener
  diferente distribución de letras)
- Selección interactiva del archivo Excel
- Detección de estudiantes repitentes/rezagados
- Exportación de resultados a PDF o Excel
"""

import json
import re
import os
import sys
import pandas as pd
from collections import defaultdict
from datetime import datetime

# ============================================================
# 1. CONFIGURACIÓN DE GRUPOS POR SEMESTRE
# ============================================================
# Cada semestre tiene su propia configuración de letras.
# Formato por línea: NOMBRE(LETRA_PRINCIPAL)=(LETRA_OVERFLOW)
#
# M=Mañana, T=Tarde, N=Noche
# El número después de M/T/N identifica el grupo dentro del turno
#
# Si un semestre no tiene configuración, se usa cada letra como
# grupo individual (comportamiento por defecto).
#
# INSTRUCCIONES PARA MODIFICAR:
# - Edita el diccionario CONFIG_GRUPOS_POR_SEMESTRE
# - Agrega o modifica la clave del semestre (1, 2, 3, ...)
# - Escribe la configuración de letras como texto multilínea

CONFIG_GRUPOS_POR_SEMESTRE = {
    1: """
M1(A)=(Q)
M2(D)=(R)
M3(E)=(V)
M4(H)=(W)
M5(I)
M6(L)
M7(M)
T1(B)=(O)
T2(F)=(P)
T3(K)=(U)
N1(C)=(T)
N2(G)=(S)
N3(J)=(X)
""",
    2: """
M1(A)
M2(D)
M3(E)
M4(H)
T1(B)
T2(F)
N1(C)
N2(G)
""",
    3: """
M1(A)=(Q)
M2(D)
M3(E)
M4(H)
M5(I)
M6(L)
M7(M)
T1(B)=(O)
T2(F)=(P)
T3(K)
N1(C)
N2(G)
N3(J)=(N)
""",
    4: """
M1(A)
M2(D)
M3(E)
T1(B)
T2(F)
N1(C)
""",
    5: """
M1(A)
M2(D)
M3(E)
M4(H)
T1(B)
T2(F)
N1(C)
N2(G)
""",
    6: """
M1(A)
T1(B)
N1(C)
""",
    7: """
M1(A)
M2(D)
M3(E)
M5(I)
M6(L)
M7(M)
T1(B)
T2(F)
T3(K)
N1(C)
N2(G)
N3(J)=(N)
""",
}

TURNOS = {
    'M': 'Mañana',
    'T': 'Tarde',
    'N': 'Noche'
}


def parsear_config_grupos(config_texto):
    """
    Parsea UNA configuración de grupos (para un semestre).
    Retorna:
    - letra_a_grupo: {letra -> nombre_grupo}
    - grupo_a_letras: {nombre_grupo -> [letras]}
    - grupo_info: {nombre_grupo -> {turno, turno_nombre, ...}}
    """
    letra_a_grupo = {}
    grupo_a_letras = {}
    grupo_info = {}

    for linea in config_texto.strip().split('\n'):
        linea = linea.strip()
        if not linea:
            continue

        match = re.match(r'^([A-Z]\d+)', linea)
        if not match:
            continue

        nombre_grupo = match.group(1)
        turno_letra = nombre_grupo[0]
        turno_numero = nombre_grupo[1:]

        letras = re.findall(r'\(([A-Z])\)', linea)
        if not letras:
            continue

        letra_principal = letras[0]
        letras_overflow = letras[1:]

        grupo_a_letras[nombre_grupo] = letras
        grupo_info[nombre_grupo] = {
            'turno': turno_letra,
            'turno_nombre': TURNOS.get(turno_letra, turno_letra),
            'numero': turno_numero,
            'letra_principal': letra_principal,
            'letras_overflow': letras_overflow
        }

        for letra in letras:
            letra_a_grupo[letra] = nombre_grupo

    return letra_a_grupo, grupo_a_letras, grupo_info


def cargar_todas_las_configs(config_dict):
    """
    Procesa el diccionario CONFIG_GRUPOS_POR_SEMESTRE.
    Retorna un dict: {semestre -> (letra_a_grupo, grupo_a_letras, grupo_info)}
    """
    configs = {}
    for semestre, texto in config_dict.items():
        configs[semestre] = parsear_config_grupos(texto)
    return configs


# ============================================================
# 2. CARGAR LA MALLA CURRICULAR
# ============================================================

def cargar_malla(ruta_json):
    with open(ruta_json, 'r', encoding='utf-8') as f:
        content = f.read()
    content = content.replace('"CR: 2"', '"CR": 2')
    data = json.loads(content)

    malla = {}
    for sem in data['semestres']:
        for mat in sem['materias']:
            codigo = mat['codigo']
            if codigo is None:
                continue
            codigo_norm = codigo.strip().replace(' ', '-')
            malla[codigo_norm] = {
                'semestre': sem['semestre'],
                'nombre_malla': mat['nombre'],
                'creditos': mat['CR'],
                'HT': mat['HT'],
                'HP': mat['HP']
            }
    return malla


# ============================================================
# 3. PARSEAR HOJA DEL EXCEL
# ============================================================

def parsear_hoja(df):
    if df.shape[0] < 7:
        return None

    fila_materia = df.iloc[4]
    codigo_materia = None
    letra_grupo = None
    nombre_materia = None

    for col_idx in range(df.shape[1]):
        val = str(fila_materia.iloc[col_idx]).strip()
        if val == 'nan' or val == '':
            continue

        match = re.match(r'^([A-Z]{2,4}[-\s]?\d{3,4})\s+([A-Z])$', val)
        if match:
            codigo_materia = match.group(1)
            letra_grupo = match.group(2)
            codigo_materia = re.sub(r'\s+', '-', codigo_materia)
            continue

        if codigo_materia and nombre_materia is None and val != 'nan':
            nombre_materia = val.strip()

    if not codigo_materia or not letra_grupo:
        return None

    estudiantes = []
    for idx in range(7, df.shape[0]):
        fila = df.iloc[idx]
        id_est = str(fila.iloc[2]).strip() if pd.notna(fila.iloc[2]) else None
        nombre = str(fila.iloc[3]).strip() if pd.notna(fila.iloc[3]) else None

        if id_est and id_est != 'nan' and nombre and nombre != 'nan':
            estudiantes.append({'id': id_est, 'nombre': nombre})

    return {
        'codigo_materia': codigo_materia,
        'letra_grupo': letra_grupo,
        'nombre_materia': nombre_materia,
        'estudiantes': estudiantes,
        'num_estudiantes': len(estudiantes)
    }


# ============================================================
# 4. PROCESAR EXCEL
# ============================================================

def procesar_excel(ruta_excel, malla):
    xls = pd.ExcelFile(ruta_excel)
    registros = []
    materias_no_encontradas = set()

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        datos = parsear_hoja(df)
        if datos is None:
            continue

        codigo = datos['codigo_materia']
        info_malla = malla.get(codigo)

        if info_malla is None:
            codigo_alt = codigo.replace('SPC-', 'SCP-')
            info_malla = malla.get(codigo_alt)
            if info_malla:
                codigo = codigo_alt
                datos['codigo_materia'] = codigo

        if info_malla:
            datos['semestre'] = info_malla['semestre']
            datos['nombre_malla'] = info_malla['nombre_malla']
            datos['creditos'] = info_malla['creditos']
        else:
            datos['semestre'] = None
            datos['nombre_malla'] = None
            datos['creditos'] = None
            materias_no_encontradas.add(datos['codigo_materia'])

        datos['hoja'] = sheet_name
        registros.append(datos)

    return registros, materias_no_encontradas


# ============================================================
# 5. DETERMINAR SEMESTRE PRINCIPAL
# ============================================================

def determinar_semestre_principal(registros):
    estudiante_semestres = defaultdict(set)
    estudiante_nombre = {}

    for r in registros:
        if r['semestre'] is None:
            continue
        for est in r['estudiantes']:
            estudiante_semestres[est['id']].add(r['semestre'])
            estudiante_nombre[est['id']] = est['nombre']

    estudiante_sem_principal = {}
    for est_id, semestres in estudiante_semestres.items():
        estudiante_sem_principal[est_id] = max(semestres)

    return estudiante_semestres, estudiante_sem_principal, estudiante_nombre


# ============================================================
# 6. GENERAR ESTADÍSTICAS
# ============================================================

def generar_estadisticas(registros, configs_grupos):
    est_semestres, est_sem_principal, est_nombre = determinar_semestre_principal(registros)

    # A) Estudiantes únicos por semestre
    estudiantes_por_semestre = defaultdict(set)
    for r in registros:
        if r['semestre'] is None:
            continue
        for est in r['estudiantes']:
            estudiantes_por_semestre[r['semestre']].add(est['id'])

    # B) Grupos reales por semestre (usando config del semestre correspondiente)
    grupo_real_semestre = defaultdict(set)
    grupo_real_letra_semestre = defaultdict(set)

    for r in registros:
        if r['semestre'] is None:
            continue
        sem = r['semestre']
        letra = r['letra_grupo']

        # Obtener mapeo de letras para ESTE semestre
        if sem in configs_grupos:
            letra_a_grupo, _, _ = configs_grupos[sem]
        else:
            letra_a_grupo = {}

        grupo_real = letra_a_grupo.get(letra, letra)

        for est in r['estudiantes']:
            grupo_real_semestre[(sem, grupo_real)].add(est['id'])
            grupo_real_letra_semestre[(sem, grupo_real, letra)].add(est['id'])

    # C) Repitentes
    repitentes = []
    repitentes_por_semestre = defaultdict(set)

    for est_id, semestres in est_semestres.items():
        sem_principal = est_sem_principal[est_id]
        semestres_inf = {s for s in semestres if s < sem_principal}

        if semestres_inf:
            repitentes.append({
                'id': est_id,
                'nombre': est_nombre[est_id],
                'semestre_principal': sem_principal,
                'semestres_donde_repite': sorted(semestres_inf),
                'todos_los_semestres': sorted(semestres)
            })
            for s in semestres_inf:
                repitentes_por_semestre[s].add(est_id)

    repitentes.sort(key=lambda x: (-x['semestre_principal'], x['nombre']))

    # D) Total
    todos = set()
    for r in registros:
        for est in r['estudiantes']:
            todos.add(est['id'])

    return {
        'estudiantes_por_semestre': estudiantes_por_semestre,
        'grupo_real_semestre': grupo_real_semestre,
        'grupo_real_letra_semestre': grupo_real_letra_semestre,
        'total_estudiantes_unicos': len(todos),
        'repitentes': repitentes,
        'repitentes_por_semestre': repitentes_por_semestre,
        'est_semestres': est_semestres,
        'est_sem_principal': est_sem_principal,
        'est_nombre': est_nombre,
    }


# ============================================================
# 7. FUNCIONES DE ORDEN PARA GRUPOS
# ============================================================

def sort_key_grupo(nombre_grupo, grupo_info):
    info = grupo_info.get(nombre_grupo)
    if info:
        turno_order = {'M': 0, 'T': 1, 'N': 2}
        return (turno_order.get(info['turno'], 3), int(info['numero']))
    return (9, str(nombre_grupo))


# ============================================================
# 8. IMPRIMIR REPORTE EN CONSOLA
# ============================================================

def imprimir_reporte(stats, configs_grupos, materias_no_encontradas):
    print("=" * 80)
    print("  REPORTE DE MATRICULADOS - MEDICINA UPDS")
    print(f"  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 80)

    total_unicos = stats['total_estudiantes_unicos']
    total_repitentes = len(stats['repitentes'])
    suma_mat = sum(len(ids) for ids in stats['estudiantes_por_semestre'].values())

    print(f"\n  TOTAL ESTUDIANTES ÚNICOS: {total_unicos}")
    print(f"  TOTAL MATRICULACIONES (suma por semestre): {suma_mat}")
    print(f"  REPITENTES/REZAGADOS DETECTADOS: {total_repitentes}")

    # --- Por semestre ---
    print(f"\n{'─' * 80}")
    print("  ESTUDIANTES POR SEMESTRE")
    print(f"{'─' * 80}")
    print(f"  {'Semestre':<12} {'Únicos':<10} {'Regulares':<12} {'Repitentes':<12} {'% Repit.':<10}")
    print(f"  {'─'*12} {'─'*10} {'─'*12} {'─'*12} {'─'*10}")

    for sem in sorted(stats['estudiantes_por_semestre'].keys()):
        unicos = len(stats['estudiantes_por_semestre'][sem])
        rep = len(stats['repitentes_por_semestre'].get(sem, set()))
        reg = unicos - rep
        pct = (rep / unicos * 100) if unicos > 0 else 0
        print(f"  Semestre {sem:<3} {unicos:<10} {reg:<12} {rep:<12} {pct:>6.1f}%")

    # --- Detalle por semestre y grupo ---
    print(f"\n{'─' * 80}")
    print("  DETALLE POR SEMESTRE Y GRUPO (letras fusionadas por semestre)")
    print(f"{'─' * 80}")

    for sem in sorted(stats['estudiantes_por_semestre'].keys()):
        total_sem = len(stats['estudiantes_por_semestre'][sem])
        rep_sem = len(stats['repitentes_por_semestre'].get(sem, set()))

        # Config de este semestre
        if sem in configs_grupos:
            _, grupo_a_letras, grupo_info = configs_grupos[sem]
        else:
            grupo_a_letras = {}
            grupo_info = {}

        print(f"\n  ┌─ SEMESTRE {sem} ─ {total_sem} únicos ({total_sem - rep_sem} regulares + {rep_sem} repitentes)")
        print(f"  │")

        grupos_del_sem = {}
        for (s, g), ids in stats['grupo_real_semestre'].items():
            if s == sem:
                grupos_del_sem[g] = ids

        turno_actual = None
        for grupo in sorted(grupos_del_sem.keys(), key=lambda g: sort_key_grupo(g, grupo_info)):
            ids = grupos_del_sem[grupo]
            count = len(ids)
            info = grupo_info.get(grupo)

            if info:
                if info['turno'] != turno_actual:
                    turno_actual = info['turno']
                    print(f"  │  ── {info['turno_nombre']} ──")

                letras = grupo_a_letras.get(grupo, [])
                desglose = []
                for letra in letras:
                    cl = len(stats['grupo_real_letra_semestre'].get((sem, grupo, letra), set()))
                    if cl > 0:
                        desglose.append(f"{letra}:{cl}")

                rep_g = sum(1 for eid in ids if eid in stats['repitentes_por_semestre'].get(sem, set()))
                des_str = " + ".join(desglose) if len(desglose) > 1 else ""
                rep_str = f" (↻ {rep_g} repit.)" if rep_g > 0 else ""

                if des_str:
                    print(f"  │    {grupo:<6} {count:>3} estudiantes  [{des_str}]{rep_str}")
                else:
                    print(f"  │    {grupo:<6} {count:>3} estudiantes{rep_str}")
            else:
                print(f"  │    ?({grupo}){' ':<2} {count:>3} estudiantes  [sin mapeo]")

        print(f"  │")
        print(f"  └─ Total: {total_sem} únicos")

    # --- Repitentes ---
    print(f"\n{'─' * 80}")
    print(f"  ESTUDIANTES REPITENTES/REZAGADOS ({total_repitentes})")
    print(f"{'─' * 80}")

    if stats['repitentes']:
        por_sem_p = defaultdict(list)
        for r in stats['repitentes']:
            por_sem_p[r['semestre_principal']].append(r)

        for sem_p in sorted(por_sem_p.keys(), reverse=True):
            estudiantes = por_sem_p[sem_p]
            print(f"\n  Semestre principal: {sem_p} ({len(estudiantes)} repitentes)")
            print(f"  {'ID':<20} {'Nombre':<45} {'Repite en'}")
            print(f"  {'─'*20} {'─'*45} {'─'*15}")
            for est in sorted(estudiantes, key=lambda x: x['nombre']):
                rep_str = ", ".join(f"S{s}" for s in est['semestres_donde_repite'])
                print(f"  {est['id']:<20} {est['nombre'][:44]:<45} {rep_str}")

    # --- Origen repitentes ---
    if stats['repitentes_por_semestre']:
        print(f"\n{'─' * 80}")
        print("  ORIGEN DE REPITENTES POR SEMESTRE")
        print(f"{'─' * 80}")
        for sem in sorted(stats['repitentes_por_semestre'].keys()):
            ids = stats['repitentes_por_semestre'][sem]
            total_en = len(stats['estudiantes_por_semestre'].get(sem, set()))
            print(f"\n  Semestre {sem}: {len(ids)} repitentes de {total_en}")
            origenes = defaultdict(int)
            for eid in ids:
                origenes[stats['est_sem_principal'][eid]] += 1
            for s_o in sorted(origenes.keys()):
                print(f"    └─ {origenes[s_o]} vienen del semestre {s_o}")

    if materias_no_encontradas:
        print(f"\n{'─' * 80}")
        print("  CÓDIGOS NO ENCONTRADOS EN LA MALLA")
        print(f"{'─' * 80}")
        for c in sorted(materias_no_encontradas):
            print(f"    {c}")


# ============================================================
# 9. EXPORTAR A EXCEL
# ============================================================

def exportar_excel(stats, configs_grupos, ruta_salida):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Estilos ---
    titulo_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    data_font = Font(name='Arial', size=10)
    repitente_fill = PatternFill('solid', fgColor='FFE0E0')
    turno_fill = PatternFill('solid', fgColor='D6E4F0')
    turno_font = Font(name='Arial', bold=True, size=10, color='1F4E79')
    border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    center = Alignment(horizontal='center', vertical='center')

    def apply_header(ws, row, cols):
        for col_idx, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def apply_data(ws, row, cols, fill=None):
        for col_idx, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            if isinstance(val, (int, float)):
                cell.alignment = center

    # ═══════════════════════════════════════════
    # Hoja 1: Resumen por semestre
    # ═══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.cell(row=1, column=1, value="REPORTE DE MATRICULADOS - MEDICINA UPDS").font = titulo_font
    ws1.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = data_font
    ws1.cell(row=3, column=1, value=f"Total estudiantes únicos: {stats['total_estudiantes_unicos']}").font = Font(name='Arial', bold=True, size=11)
    ws1.cell(row=4, column=1, value=f"Repitentes detectados: {len(stats['repitentes'])}").font = Font(name='Arial', bold=True, size=11, color='CC0000')

    row = 6
    apply_header(ws1, row, ['Semestre', 'Únicos', 'Regulares', 'Repitentes', '% Repitentes'])
    row += 1
    for sem in sorted(stats['estudiantes_por_semestre'].keys()):
        unicos = len(stats['estudiantes_por_semestre'][sem])
        rep = len(stats['repitentes_por_semestre'].get(sem, set()))
        reg = unicos - rep
        pct = round(rep / unicos * 100, 1) if unicos > 0 else 0
        apply_data(ws1, row, [f'Semestre {sem}', unicos, reg, rep, f'{pct}%'])
        row += 1

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws1.column_dimensions[col].width = 18

    # ═══════════════════════════════════════════
    # Hoja 2: Detalle por grupo
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("Grupos por Semestre")
    ws2.cell(row=1, column=1, value="DETALLE POR GRUPO Y SEMESTRE").font = titulo_font
    row = 3

    for sem in sorted(stats['estudiantes_por_semestre'].keys()):
        if sem in configs_grupos:
            _, grupo_a_letras, grupo_info = configs_grupos[sem]
        else:
            grupo_a_letras = {}
            grupo_info = {}

        total_sem = len(stats['estudiantes_por_semestre'][sem])
        rep_sem = len(stats['repitentes_por_semestre'].get(sem, set()))
        ws2.cell(row=row, column=1, value=f"SEMESTRE {sem} — {total_sem} únicos ({total_sem - rep_sem} regulares + {rep_sem} repitentes)")
        ws2.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=11, color='1F4E79')
        row += 1

        apply_header(ws2, row, ['Grupo', 'Turno', 'Letras', 'Matriculados', 'Repitentes', 'Regulares'])
        row += 1

        grupos_del_sem = {}
        for (s, g), ids in stats['grupo_real_semestre'].items():
            if s == sem:
                grupos_del_sem[g] = ids

        for grupo in sorted(grupos_del_sem.keys(), key=lambda g: sort_key_grupo(g, grupo_info)):
            ids = grupos_del_sem[grupo]
            count = len(ids)
            info = grupo_info.get(grupo)
            rep_g = sum(1 for eid in ids if eid in stats['repitentes_por_semestre'].get(sem, set()))

            if info:
                letras = grupo_a_letras.get(grupo, [])
                letras_str = "+".join(letras)
                turno = info['turno_nombre']
            else:
                letras_str = grupo
                turno = '?'

            fill = repitente_fill if rep_g > 0 else None
            apply_data(ws2, row, [grupo, turno, letras_str, count, rep_g, count - rep_g], fill)
            row += 1

        row += 1  # Espacio entre semestres

    for col, w in [('A', 10), ('B', 12), ('C', 10), ('D', 14), ('E', 13), ('F', 12)]:
        ws2.column_dimensions[col].width = w

    # ═══════════════════════════════════════════
    # Hoja 3: Lista de repitentes
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("Repitentes")
    ws3.cell(row=1, column=1, value=f"ESTUDIANTES REPITENTES ({len(stats['repitentes'])})").font = titulo_font
    row = 3
    apply_header(ws3, row, ['ID', 'Nombre', 'Semestre Principal', 'Semestres donde repite', 'Todos los semestres'])
    row += 1

    for est in stats['repitentes']:
        rep_str = ", ".join(f"S{s}" for s in est['semestres_donde_repite'])
        all_str = ", ".join(f"S{s}" for s in est['todos_los_semestres'])
        apply_data(ws3, row, [est['id'], est['nombre'], est['semestre_principal'], rep_str, all_str], repitente_fill)
        row += 1

    for col, w in [('A', 22), ('B', 48), ('C', 20), ('D', 28), ('E', 22)]:
        ws3.column_dimensions[col].width = w

    # ═══════════════════════════════════════════
    # Hoja 4: Origen de repitentes
    # ═══════════════════════════════════════════
    ws4 = wb.create_sheet("Origen Repitentes")
    ws4.cell(row=1, column=1, value="ORIGEN DE REPITENTES POR SEMESTRE").font = titulo_font
    row = 3
    apply_header(ws4, row, ['Semestre donde repite', 'Total repitentes', 'Vienen del semestre', 'Cantidad'])
    row += 1

    for sem in sorted(stats['repitentes_por_semestre'].keys()):
        ids = stats['repitentes_por_semestre'][sem]
        origenes = defaultdict(int)
        for eid in ids:
            origenes[stats['est_sem_principal'][eid]] += 1
        first = True
        for s_o in sorted(origenes.keys()):
            if first:
                apply_data(ws4, row, [f'Semestre {sem}', len(ids), f'Semestre {s_o}', origenes[s_o]])
                first = False
            else:
                apply_data(ws4, row, ['', '', f'Semestre {s_o}', origenes[s_o]])
            row += 1
        row += 1

    for col, w in [('A', 22), ('B', 18), ('C', 22), ('D', 12)]:
        ws4.column_dimensions[col].width = w

    wb.save(ruta_salida)
    print(f"\n  ✓ Archivo Excel exportado: {ruta_salida}")


# ============================================================
# 10. EXPORTAR A PDF
# ============================================================

def exportar_pdf(stats, configs_grupos, ruta_salida):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.units import inch

    doc = SimpleDocTemplate(ruta_salida, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TituloReporte', fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor('#1F4E79'), spaceAfter=6))
    styles.add(ParagraphStyle(name='Subtitulo', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#1F4E79'), spaceAfter=4, spaceBefore=12))
    styles.add(ParagraphStyle(name='Info', fontName='Helvetica', fontSize=10, spaceAfter=2))
    styles.add(ParagraphStyle(name='CeldaPeq', fontName='Helvetica', fontSize=8))

    story = []
    AZUL = colors.HexColor('#1F4E79')
    GRIS_CLARO = colors.HexColor('#F2F2F2')
    ROJO_CLARO = colors.HexColor('#FFE0E0')

    header_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), AZUL),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])

    # --- Portada ---
    story.append(Paragraph("REPORTE DE MATRICULADOS", styles['TituloReporte']))
    story.append(Paragraph("Carrera de Medicina - UPDS", styles['Subtitulo']))
    story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Info']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Total estudiantes únicos: <b>{stats['total_estudiantes_unicos']}</b>", styles['Info']))
    story.append(Paragraph(f"Repitentes detectados: <b>{len(stats['repitentes'])}</b>", styles['Info']))
    story.append(Spacer(1, 20))

    # --- Tabla resumen ---
    story.append(Paragraph("RESUMEN POR SEMESTRE", styles['Subtitulo']))
    data = [['Semestre', 'Únicos', 'Regulares', 'Repitentes', '% Repitentes']]
    for sem in sorted(stats['estudiantes_por_semestre'].keys()):
        u = len(stats['estudiantes_por_semestre'][sem])
        r = len(stats['repitentes_por_semestre'].get(sem, set()))
        pct = f"{r/u*100:.1f}%" if u > 0 else "0%"
        data.append([f'Semestre {sem}', str(u), str(u-r), str(r), pct])

    t = Table(data, colWidths=[100, 80, 80, 80, 90])
    t.setStyle(header_style)
    story.append(t)

    # --- Detalle grupos por semestre ---
    for sem in sorted(stats['estudiantes_por_semestre'].keys()):
        story.append(PageBreak())
        total_sem = len(stats['estudiantes_por_semestre'][sem])
        rep_sem = len(stats['repitentes_por_semestre'].get(sem, set()))
        story.append(Paragraph(f"SEMESTRE {sem} — {total_sem} únicos ({total_sem - rep_sem} reg. + {rep_sem} repit.)", styles['Subtitulo']))

        if sem in configs_grupos:
            _, grupo_a_letras, grupo_info = configs_grupos[sem]
        else:
            grupo_a_letras = {}
            grupo_info = {}

        data = [['Grupo', 'Turno', 'Letras', 'Matriculados', 'Repitentes', 'Regulares']]
        row_fills = []

        grupos_del_sem = {}
        for (s, g), ids in stats['grupo_real_semestre'].items():
            if s == sem:
                grupos_del_sem[g] = ids

        for grupo in sorted(grupos_del_sem.keys(), key=lambda g: sort_key_grupo(g, grupo_info)):
            ids = grupos_del_sem[grupo]
            count = len(ids)
            info = grupo_info.get(grupo)
            rep_g = sum(1 for eid in ids if eid in stats['repitentes_por_semestre'].get(sem, set()))

            turno = info['turno_nombre'] if info else '?'
            letras_str = "+".join(grupo_a_letras.get(grupo, [grupo]))
            data.append([grupo, turno, letras_str, str(count), str(rep_g), str(count - rep_g)])
            row_fills.append(rep_g > 0)

        t = Table(data, colWidths=[60, 80, 60, 90, 80, 80])
        style_cmds = list(header_style.getCommands())
        for i, has_rep in enumerate(row_fills):
            if has_rep:
                style_cmds.append(('BACKGROUND', (0, i+1), (-1, i+1), ROJO_CLARO))
        t.setStyle(TableStyle(style_cmds))
        story.append(t)

    # --- Lista de repitentes ---
    story.append(PageBreak())
    story.append(Paragraph(f"ESTUDIANTES REPITENTES ({len(stats['repitentes'])})", styles['Subtitulo']))

    if stats['repitentes']:
        data = [['ID', 'Nombre', 'Sem. Principal', 'Repite en']]
        for est in stats['repitentes']:
            rep_str = ", ".join(f"S{s}" for s in est['semestres_donde_repite'])
            nombre = est['nombre'][:50]
            data.append([est['id'], nombre, str(est['semestre_principal']), rep_str])

        t = Table(data, colWidths=[130, 250, 90, 100])
        t.setStyle(header_style)
        story.append(t)

    doc.build(story)
    print(f"\n  ✓ Archivo PDF exportado: {ruta_salida}")


# ============================================================
# 11. SELECCIÓN DE ARCHIVO
# ============================================================

def seleccionar_archivo():
    """Permite al usuario seleccionar el archivo Excel a analizar."""
    print("\n" + "=" * 60)
    print("  SISTEMA DE ANÁLISIS DE MATRICULADOS")
    print("  Medicina - UPDS")
    print("=" * 60)

    while True:
        print("\n  Ingresa la ruta completa del archivo Excel (.xlsx)")
        print("  (o escribe 'salir' para terminar)")
        ruta = input("\n  Ruta: ").strip()

        if ruta.lower() == 'salir':
            sys.exit(0)

        # Limpiar comillas si las pegó
        ruta = ruta.strip('"').strip("'")

        if not os.path.exists(ruta):
            print(f"\n  ✗ No se encontró el archivo: {ruta}")
            continue

        if not ruta.lower().endswith(('.xlsx', '.xls')):
            print(f"\n  ✗ El archivo debe ser .xlsx o .xls")
            continue

        print(f"\n  ✓ Archivo encontrado: {os.path.basename(ruta)}")
        print(f"    Tamaño: {os.path.getsize(ruta) / 1024:.1f} KB")
        return ruta


def preguntar_exportacion(nombre_base):
    """Pregunta al usuario si desea exportar los resultados."""
    print(f"\n{'─' * 60}")
    print("  ¿Deseas exportar los resultados?")
    print("  1) Exportar a Excel (.xlsx)")
    print("  2) Exportar a PDF (.pdf)")
    print("  3) Exportar ambos (Excel + PDF)")
    print("  4) No exportar, solo ver en consola")

    while True:
        opcion = input("\n  Opción [1/2/3/4]: ").strip()
        if opcion in ('1', '2', '3', '4'):
            break
        print("  Opción no válida. Ingresa 1, 2, 3 o 4.")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    resultados = []

    if opcion in ('1', '3'):
        resultados.append(('excel', f'Reporte_Matriculados_{timestamp}.xlsx'))
    if opcion in ('2', '3'):
        resultados.append(('pdf', f'Reporte_Matriculados_{timestamp}.pdf'))

    return resultados


# ============================================================
# 12. EJECUCIÓN PRINCIPAL
# ============================================================

if __name__ == '__main__':

    # Determinar ruta de malla (buscar en el mismo directorio del script)
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    RUTA_MALLA = os.path.join(SCRIPT_DIR, 'Malla-curricular-medicina-UPDS.txt')

    # Si no la encuentra, preguntar
    if not os.path.exists(RUTA_MALLA):
        print("\n  No se encontró 'Malla-curricular-medicina-UPDS.txt' en la carpeta del script.")
        print("  Ingresa la ruta de la malla curricular:")
        RUTA_MALLA = input("  Ruta: ").strip().strip('"').strip("'")

    if not os.path.exists(RUTA_MALLA):
        print(f"\n  ✗ No se encontró: {RUTA_MALLA}")
        sys.exit(1)

    # ── Seleccionar archivo Excel ──
    RUTA_EXCEL = seleccionar_archivo()

    # ── Cargar configuraciones ──
    print("\n  Cargando configuraciones de grupos por semestre...")
    configs_grupos = cargar_todas_las_configs(CONFIG_GRUPOS_POR_SEMESTRE)
    for sem in sorted(configs_grupos.keys()):
        _, g_a_l, g_info = configs_grupos[sem]
        n_grupos = len(g_info)
        n_letras = sum(len(v) for v in g_a_l.values())
        print(f"    Semestre {sem}: {n_grupos} grupos, {n_letras} letras mapeadas")

    print("\n  Cargando malla curricular...")
    malla = cargar_malla(RUTA_MALLA)
    print(f"    {len(malla)} materias en la malla")

    print("\n  Procesando archivo Excel...")
    registros, no_encontradas = procesar_excel(RUTA_EXCEL, malla)
    print(f"    {len(registros)} registros procesados")

    print("\n  Generando estadísticas...")
    stats = generar_estadisticas(registros, configs_grupos)

    # ── Mostrar reporte en consola ──
    imprimir_reporte(stats, configs_grupos, no_encontradas)

    # ── Preguntar exportación ──
    nombre_base = os.path.splitext(os.path.basename(RUTA_EXCEL))[0]
    exportaciones = preguntar_exportacion(nombre_base)

    # Directorio de salida = misma carpeta del Excel
    dir_salida = os.path.dirname(os.path.abspath(RUTA_EXCEL))

    for tipo, nombre in exportaciones:
        ruta_out = os.path.join(dir_salida, nombre)
        try:
            if tipo == 'excel':
                exportar_excel(stats, configs_grupos, ruta_out)
            elif tipo == 'pdf':
                exportar_pdf(stats, configs_grupos, ruta_out)
        except Exception as e:
            print(f"\n  ✗ Error al exportar {tipo}: {e}")
            print(f"    Asegúrate de tener instaladas las dependencias.")

    print(f"\n{'═' * 60}")
    print("  Análisis completado.")
    print(f"{'═' * 60}\n")
